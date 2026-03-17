"""viewer_server.py – T24 3D live viewer/editor
Run: python viewer_server.py  →  http://localhost:5000
"""
import sys, os, time, shutil, subprocess, threading, webbrowser, json

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# Directories to search for source _t24.json files
# Defaults to SCRIPT_DIR; user can add paths via --search-dir flag or T24_SEARCH_DIRS env var
_SEARCH_DIRS = [SCRIPT_DIR]
_env_dirs = os.environ.get("T24_SEARCH_DIRS", "")
if _env_dirs:
    for d in _env_dirs.split(os.pathsep):
        if d.strip() and os.path.isdir(d.strip()):
            _SEARCH_DIRS.append(d.strip())

# Also accept --search-dir flags from command line
_i = 1
while _i < len(sys.argv):
    if sys.argv[_i] == "--search-dir" and _i + 1 < len(sys.argv):
        _d = sys.argv[_i + 1]
        if os.path.isdir(_d) and _d not in _SEARCH_DIRS:
            _SEARCH_DIRS.append(_d)
        _i += 2
    else:
        _i += 1

try:
    from flask import Flask, jsonify, request, send_from_directory
    import openpyxl
except ImportError:
    subprocess.run([sys.executable, "-m", "pip", "install", "flask", "openpyxl"], check=True)
    from flask import Flask, jsonify, request, send_from_directory
    import openpyxl

from t24_utils import parse_azimuth

app = Flask(__name__)

def azimuth(s):
    return parse_azimuth(s)

def flt(v, default=0.0):
    try: return float(v) if v not in (None, "") else default
    except: return default

# Maps filename → directory it was found in
# _BROWSED_DIRS takes priority — set when user browses to a file
_FILE_DIRS = {}
_BROWSED_DIRS = {}

def find_xlsx():
    """Find all takeoff xlsx files across SCRIPT_DIR and _SEARCH_DIRS.
    Returns list of filenames."""
    seen = {}
    for d in [SCRIPT_DIR] + _SEARCH_DIRS:
        if not os.path.isdir(d):
            continue
        for f in os.listdir(d):
            if f.lower().endswith(".xlsx") and not f.startswith("~$") and f not in seen:
                seen[f] = d
    _FILE_DIRS.update(seen)
    # Browsed dirs always win
    _FILE_DIRS.update(_BROWSED_DIRS)
    return sorted(seen.keys())

def resolve_path(fn):
    """Resolve a filename to its full path, checking _FILE_DIRS then SCRIPT_DIR."""
    if fn in _FILE_DIRS:
        return os.path.join(_FILE_DIRS[fn], fn)
    return os.path.join(SCRIPT_DIR, fn)

def read_data(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    p = wb["Project"]
    project = {
        "name": str(p.cell(2,2).value or ""),
        "address": str(p.cell(3,2).value or ""),
        "climate_zone": str(p.cell(4,2).value or ""),
        "building_type": str(p.cell(5,2).value or "MultiFamily"),
        "front_orientation": flt(p.cell(6,2).value, 0),
        "standards_version": str(p.cell(7,2).value or "2022"),
    }
    zones = []
    for i, row in enumerate(wb["Zones"].iter_rows(min_row=2, values_only=True), 2):
        if not row[0] or str(row[0]).startswith("#"): continue
        zid = str(row[0]).strip()
        zones.append({"_row":i,"id":zid,"name":str(row[1] or zid),
                       "area":flt(row[2]),"height":flt(row[3],9.0),
                       "cond_type":str(row[4] or "Conditioned"),
                       "occ_type":str(row[5] or "") if len(row)>5 else "",
                       "floor":int(flt(row[6],1)) if len(row)>6 and row[6] not in (None,"") else 1})
    walls = []
    for i, row in enumerate(wb["Walls"].iter_rows(min_row=2, values_only=True), 2):
        if not row[0] or str(row[0]).startswith("#"): continue
        wid = str(row[0]).strip()
        walls.append({"_row":i,"id":wid,"zone_id":str(row[1] or "").strip(),
                       "name":str(row[2] or wid),"type":str(row[3] or "Exterior Wall"),
                       "orientation":str(row[4] or ""),"azimuth":azimuth(row[4]),
                       "area":flt(row[5]),"construction":str(row[6] or ""),
                       "adj_zone":str(row[7] or "").strip() if len(row)>7 else ""})
    openings = []
    for i, row in enumerate(wb["Openings"].iter_rows(min_row=2, values_only=True), 2):
        if not row[0] or str(row[0]).startswith("#"): continue
        oid = str(row[0]).strip()
        openings.append({"_row":i,"id":oid,"wall_id":str(row[1] or "").strip(),
                          "name":str(row[2] or oid),"type":str(row[3] or "Window"),
                          "area":flt(row[4]),
                          "ufactor":flt(row[5],None) if row[5] not in (None,"") else None,
                          "shgc":flt(row[6],None) if len(row)>6 and row[6] not in (None,"") else None})
    return {"project":project,"zones":zones,"walls":walls,"openings":openings}

def coerce(v):
    if v in (None, ""): return None
    try: return int(v) if "." not in str(v) else float(v)
    except: return v

@app.route("/")
def index(): return send_from_directory(SCRIPT_DIR, "viewer.html")

@app.route("/api/files")
def api_files(): return jsonify(find_xlsx())

@app.route("/api/data")
def api_data():
    fn = request.args.get("file") or (find_xlsx() or [""])[0]
    path = resolve_path(fn)
    if not os.path.exists(path): return jsonify({"error": f"Not found: {fn}"}), 404
    # Re-run JSON→Excel conversion if source JSON is newer
    maybe_reconvert(path)
    try:
        d = read_data(path)
        d["_file"] = fn
        d["_mtime"] = os.path.getmtime(path)
        return jsonify(d)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

def find_source_json(xlsx_path):
    """Find the _t24.json that produced this _t24_takeoff.xlsx."""
    base = os.path.splitext(xlsx_path)[0]            # …_t24_takeoff
    if base.lower().endswith("_takeoff"):
        json_path = base[:-len("_takeoff")] + ".json" # …_t24.json
        if os.path.exists(json_path):
            return json_path
    # Also search all additional working directories for the JSON
    stem = os.path.basename(base)
    if stem.lower().endswith("_takeoff"):
        json_name = stem[:-len("_takeoff")] + ".json"
        # Search in all directories the server knows about
        for d in _SEARCH_DIRS:
            candidate = os.path.join(d, json_name)
            if os.path.exists(candidate):
                return candidate
    return None

def maybe_reconvert(xlsx_path):
    """If source _t24.json is newer than the xlsx, re-run tz_to_excel.py
    and copy results back into SCRIPT_DIR if they were written elsewhere."""
    json_path = find_source_json(xlsx_path)
    if not json_path:
        return False
    try:
        if os.path.getmtime(json_path) > os.path.getmtime(xlsx_path):
            tz_script = os.path.join(SCRIPT_DIR, "tz_to_excel.py")
            if not os.path.exists(tz_script):
                return False
            subprocess.run(
                [sys.executable, tz_script, json_path],
                capture_output=True, text=True, timeout=30)
            # tz_to_excel.py writes output next to the JSON file.
            # If that's a different directory, copy results into SCRIPT_DIR.
            json_dir = os.path.dirname(os.path.abspath(json_path))
            if os.path.normcase(json_dir) != os.path.normcase(SCRIPT_DIR):
                xlsx_name = os.path.basename(xlsx_path)
                geo_name = os.path.splitext(xlsx_name)[0] + "_geometry.json"
                for fname in (xlsx_name, geo_name):
                    src = os.path.join(json_dir, fname)
                    if os.path.exists(src):
                        shutil.copy2(src, os.path.join(SCRIPT_DIR, fname))
            return True
    except Exception:
        pass
    return False

@app.route("/api/mtime")
def api_mtime():
    fn = request.args.get("file","")
    p = resolve_path(fn)
    if os.path.exists(p):
        maybe_reconvert(p)
    return jsonify({"mtime": os.path.getmtime(p) if os.path.exists(p) else 0})

@app.route("/api/update", methods=["POST"])
def api_update():
    d = request.json
    path = resolve_path(d["file"])
    try:
        wb = openpyxl.load_workbook(path)
        wb[d["sheet"]].cell(int(d["row"]), int(d["col"]), coerce(d["value"]))
        wb.save(path)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/add_row", methods=["POST"])
def api_add_row():
    d = request.json
    path = resolve_path(d["file"])
    try:
        wb = openpyxl.load_workbook(path)
        ws = wb[d["sheet"]]
        # Find last non-empty, non-comment data row
        last = 1
        for r in range(2, ws.max_row + 1):
            v = ws.cell(r, 1).value
            if v and not str(v).startswith("#"): last = r
        new_row = last + 1
        ws.insert_rows(new_row)
        for j, val in enumerate(d.get("values", []), 1):
            if val not in (None, ""): ws.cell(new_row, j, coerce(val))
        wb.save(path)
        return jsonify({"ok": True, "row": new_row})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/delete_row", methods=["POST"])
def api_delete_row():
    d = request.json
    path = resolve_path(d["file"])
    try:
        wb = openpyxl.load_workbook(path)
        wb[d["sheet"]].delete_rows(int(d["row"]))
        wb.save(path)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/geometry")
def api_geometry():
    fn = request.args.get("file", "")
    path = resolve_path(fn)
    maybe_reconvert(path)
    base = os.path.splitext(path)[0]
    geo_path = base + "_geometry.json"
    if not os.path.exists(geo_path):
        return jsonify(None)
    try:
        import json
        with open(geo_path) as f:
            return jsonify(json.load(f))
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/browse")
def api_browse():
    """Open a file dialog to pick an xlsx file. Registers its directory for future lookups."""
    try:
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk()
        root.withdraw()
        root.attributes('-topmost', True)
        path = filedialog.askopenfilename(
            title="Open T24 Takeoff Excel",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        root.destroy()
        if not path:
            return jsonify({"file": None})
        fn = os.path.basename(path)
        d = os.path.dirname(os.path.abspath(path))
        # Register the directory so resolve_path can find it (browsed wins over auto-found)
        _FILE_DIRS[fn] = d
        _BROWSED_DIRS[fn] = d
        if d not in _SEARCH_DIRS:
            _SEARCH_DIRS.append(d)
        # Auto-convert if there's a newer source JSON
        maybe_reconvert(path)
        return jsonify({"file": fn})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/browse-jsons")
def api_browse_jsons():
    """Open a multi-file dialog to pick JSON files for merging."""
    try:
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk()
        root.withdraw()
        root.attributes('-topmost', True)
        paths = filedialog.askopenfilenames(
            title="Select T24 JSON files to merge",
            filetypes=[("T24 JSON", "*_t24.json"), ("JSON files", "*.json"), ("All files", "*.*")])
        root.destroy()
        if not paths:
            return jsonify({"files": []})
        return jsonify({"files": list(paths)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/import-json", methods=["POST"])
def api_import_json():
    """Import JSON file(s), apply floor override, convert to Excel."""
    json_paths = request.json.get("files", [])
    floor = request.json.get("floor", 1)
    if not json_paths:
        return jsonify({"error": "No files provided"}), 400
    for p in json_paths:
        if not os.path.exists(p):
            return jsonify({"error": f"File not found: {p}"}), 400
    try:
        # Apply floor override to each JSON before converting
        import tempfile
        patched_paths = []
        for jp in json_paths:
            with open(jp, 'r', encoding='utf-8') as f:
                data = json.load(f)
            # Override floor on all zones
            for z in data.get('zones', []):
                z['floor'] = floor
            # Write patched version to temp file next to original
            out_dir = os.path.dirname(os.path.abspath(jp))
            base = os.path.splitext(os.path.basename(jp))[0]
            patched = os.path.join(out_dir, base + f"_fl{floor}.json")
            with open(patched, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2)
            patched_paths.append(patched)

        tz_script = os.path.join(SCRIPT_DIR, "tz_to_excel.py")
        r = subprocess.run(
            [sys.executable, tz_script] + patched_paths,
            capture_output=True, text=True, timeout=60)

        # Clean up temp files
        for pp in patched_paths:
            try: os.remove(pp)
            except: pass

        if r.returncode != 0:
            return jsonify({"error": f"Conversion failed:\n{r.stderr}"}), 500

        # Figure out the output filename
        if len(patched_paths) == 1:
            base = os.path.splitext(os.path.basename(patched_paths[0]))[0].replace('_t24', '')
        else:
            base = "t24_combined"
        out_dir = os.path.dirname(os.path.abspath(json_paths[0]))
        xlsx_name = base + "_t24_takeoff.xlsx"
        xlsx_path = os.path.join(out_dir, xlsx_name)

        if os.path.exists(xlsx_path):
            d = os.path.dirname(os.path.abspath(xlsx_path))
            fn = os.path.basename(xlsx_path)
            _FILE_DIRS[fn] = d
            _BROWSED_DIRS[fn] = d
            if d not in _SEARCH_DIRS:
                _SEARCH_DIRS.append(d)
            return jsonify({"ok": True, "file": fn, "count": len(json_paths)})
        else:
            return jsonify({"error": "Excel not created.\n" + r.stdout + r.stderr}), 500
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/generate", methods=["POST"])
def api_generate():
    fn = request.json.get("file")
    xlsx = resolve_path(fn)
    # Always put output XML next to the xlsx file
    xlsx_dir = os.path.dirname(os.path.abspath(xlsx))
    out = os.path.join(xlsx_dir, os.path.splitext(os.path.basename(xlsx))[0] + ".xml")
    try:
        r = subprocess.run(
            [sys.executable, os.path.join(SCRIPT_DIR, "generate_gbxml.py"), xlsx, out],
            capture_output=True, text=True, timeout=30)
        return jsonify({"ok": r.returncode==0, "output": r.stdout+r.stderr,
                        "outfile": out})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

if __name__ == "__main__":
    print("="*50)
    print("  T24 3D Viewer  ->  http://localhost:5000")
    print("  Ctrl+C to stop")
    print("="*50)
    threading.Thread(
        target=lambda: (time.sleep(1.2), webbrowser.open("http://localhost:5000")),
        daemon=True).start()
    app.run(host="127.0.0.1", port=5000, debug=False, use_reloader=False)
