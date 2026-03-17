"""
tz_to_excel.py  –  T24 TakeOff Wizard: JSON → Excel + geometry sidecar

Usage:
    python tz_to_excel.py "path/to/drawing_t24.json"
    python tz_to_excel.py file1.json file2.json file3.json   (merge all into one Excel)

Outputs (same folder as first input):
    drawing_t24_takeoff.xlsx   – ready for generate_gbxml.py
    drawing_t24_geometry.json  – real XY coordinates for the 3D viewer

How it works:
    1. Reads zone XDATA (vertices, ceiling height, floor, condition, occupancy)
    2. Reads manually-placed wall markers from JSON (placed by TZ-WALL command)
    3. Adds 4 horizontal surfaces per zone (slab, floor, ceiling, roof)
    4. Associates windows/doors with their nearest wall segment
    5. Associates skylights with their zone (point-in-polygon)
    6. Writes the Excel template and a geometry JSON sidecar
    7. Multiple JSONs: merges zones/walls/openings from all files
"""

import sys, os, json, math, glob
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from t24_utils import az_to_cardinal

# ── Constants ─────────────────────────────────────────────────────────────────
IN_PER_FT = 12.0          # Drawing units are inches
OPENING_ASSIGN_TOL = 120.0 # Inches – max distance to assign opening to a wall

# ── Geometry helpers ──────────────────────────────────────────────────────────

def in2ft(v):
    return v / IN_PER_FT

def pts_in_ft(pts):
    return [[in2ft(p[0]), in2ft(p[1])] for p in pts]

def seg_length_in(p1, p2):
    return math.sqrt((p2[0]-p1[0])**2 + (p2[1]-p1[1])**2)

def seg_length_ft(p1, p2):
    return seg_length_in(p1, p2) / IN_PER_FT

def point_in_polygon(pt, poly):
    x, y = pt[0], pt[1]
    n = len(poly)
    inside = False
    j = n - 1
    for i in range(n):
        xi, yi = poly[i]
        xj, yj = poly[j]
        if ((yi > y) != (yj > y)) and (x < (xj - xi) * (y - yi) / (yj - yi) + xi):
            inside = not inside
        j = i
    return inside

def pt_to_seg_dist(pt, p1, p2):
    """Perpendicular distance from pt to infinite line through p1→p2, and projection param t."""
    dx = p2[0] - p1[0]
    dy = p2[1] - p1[1]
    if dx == 0 and dy == 0:
        return math.sqrt((pt[0]-p1[0])**2 + (pt[1]-p1[1])**2), 0.5
    t = ((pt[0]-p1[0])*dx + (pt[1]-p1[1])*dy) / (dx*dx + dy*dy)
    t = max(0.0, min(1.0, t))
    cx = p1[0] + t * dx
    cy = p1[1] + t * dy
    return math.sqrt((pt[0]-cx)**2 + (pt[1]-cy)**2), t

# ── Main processing ───────────────────────────────────────────────────────────

def process(json_path):
    with open(json_path, 'r', encoding='utf-8') as f:
        raw = f.read()
    data = json.loads(raw)

    zones    = data.get('zones', [])
    openings = data.get('openings', [])
    raw_walls = data.get('walls', [])
    climate_zone    = data.get('climate_zone', '')
    front_orient    = data.get('front_orientation', 'South')

    # ── Read wall markers placed by TZ-WALL command ───────────────────────
    all_walls = []
    for w in raw_walls:
        wid      = w.get('id', 'W-???')
        wtype    = w.get('type', 'Exterior Wall')
        zone_id  = w.get('zone_id', '')
        az       = w.get('azimuth', 0.0)
        area     = w.get('area_sqft', 0.0)
        az_name  = az_to_name(az)

        all_walls.append({
            'zone_id':     zone_id,
            'wall_id':     wid,
            'name':        f"{az_name} Wall",
            'type':        wtype,
            'orientation': az_to_cardinal(az) if 'exterior' in wtype.lower() else '',
            'azimuth':     az,
            'area_sqft':   round(area, 1),
            'adj_zone':    '',
            'p1':          w.get('p1'),
            'p2':          w.get('p2'),
        })

    # ── 4 horizontal surfaces per zone ─────────────────────────────────────
    horiz_rows = []
    for zone in zones:
        zid = zone['id']
        area = zone.get('area_sqft', 0)
        horiz_rows.append({
            'zone_id': zid, 'wall_id': f"{zid}-SLAB",
            'name': 'Slab on Grade', 'type': 'Slab on Grade', 'orientation': '',
            'azimuth': 0, 'area_sqft': round(area, 1), 'adj_zone': '', 'p1': None, 'p2': None,
        })
        horiz_rows.append({
            'zone_id': zid, 'wall_id': f"{zid}-FLOOR",
            'name': 'Interior Floor', 'type': 'Raised Floor', 'orientation': '',
            'azimuth': 0, 'area_sqft': round(area, 1), 'adj_zone': '', 'p1': None, 'p2': None,
        })
        horiz_rows.append({
            'zone_id': zid, 'wall_id': f"{zid}-CEIL",
            'name': 'Interior Ceiling', 'type': 'Interior Ceiling', 'orientation': '',
            'azimuth': 0, 'area_sqft': round(area, 1), 'adj_zone': '', 'p1': None, 'p2': None,
        })
        horiz_rows.append({
            'zone_id': zid, 'wall_id': f"{zid}-ROOF",
            'name': 'Roof', 'type': 'Roof', 'orientation': '',
            'azimuth': 0, 'area_sqft': round(area, 1), 'adj_zone': '', 'p1': None, 'p2': None,
        })

    all_walls.extend(horiz_rows)

    # ── Assign openings to zones and walls ────────────────────────────────
    assigned_openings = []
    for o in openings:
        otype   = o.get('type', 'Window')
        pos     = o.get('position', [0, 0])   # raw inches
        area    = o.get('area_sqft', 0)
        uf      = o.get('u_factor', 0)
        shgc    = o.get('shgc', 0)
        oid     = o.get('id', 'O-???')

        # Find containing zone (point-in-polygon, with nearest-zone fallback
        # for openings placed on or just outside the boundary edge)
        host_zone = None
        for zone in zones:
            if point_in_polygon(pos, zone['vertices']):
                host_zone = zone['id']
                break

        if not host_zone:
            # Fallback: find nearest zone boundary (openings on edges fail PIP)
            best_zone_dist = float('inf')
            for zone in zones:
                verts = zone['vertices']
                for k in range(len(verts)):
                    p1 = verts[k]
                    p2 = verts[(k + 1) % len(verts)]
                    dist, _ = pt_to_seg_dist(pos, p1, p2)
                    if dist < best_zone_dist:
                        best_zone_dist = dist
                        host_zone = zone['id']
            if host_zone and best_zone_dist < OPENING_ASSIGN_TOL:
                print(f"  NOTE: {otype} {oid} on zone edge — assigned to {host_zone} (dist={best_zone_dist/12:.1f}ft)")
            else:
                host_zone = None

        if not host_zone:
            print(f"  WARNING: {otype} {oid} is not inside any zone boundary — marked UNASSIGNED")
            wall_id = 'UNASSIGNED'
        elif otype == 'Skylight':
            # Skylight → find roof wall of the zone
            wall_id = f"{host_zone}-ROOF"
        else:
            # Window / Door → find nearest vertical wall segment in same zone
            best_wall = None
            best_dist = float('inf')
            for w in all_walls:
                if w['p1'] is None:   # skip horiz placeholders
                    continue
                if w['zone_id'] != host_zone:
                    continue
                dist, _ = pt_to_seg_dist(pos, w['p1'], w['p2'])
                if dist < best_dist:
                    best_dist = dist
                    best_wall = w['wall_id']
            wall_id = best_wall or 'UNASSIGNED'

        assigned_openings.append({
            'id':      oid,
            'wall_id': wall_id,
            'name':    f"{otype} {oid}",
            'type':    otype,
            'area_sqft': area,
            'ufactor': uf if uf else None,
            'shgc':    shgc if shgc else None,
            'position': pos,
        })

    return zones, all_walls, assigned_openings, climate_zone, front_orient


def az_to_name(az):
    dirs = [
        (0,   'North'), (45,  'NE'), (90,  'East'),  (135, 'SE'),
        (180, 'South'), (225, 'SW'), (270, 'West'),  (315, 'NW')
    ]
    az = az % 360
    return min(dirs, key=lambda d: abs(((az - d[0] + 180) % 360) - 180))[1]


# ── Excel writer ──────────────────────────────────────────────────────────────

def write_excel(json_path, zones, walls, openings, climate_zone='', front_orient='South'):
    out_dir  = os.path.dirname(json_path)
    base     = os.path.splitext(os.path.basename(json_path))[0].replace('_t24', '')
    out_path = os.path.join(out_dir, base + '_t24_takeoff.xlsx')

    wb = Workbook()
    HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
    HDR_FONT  = Font(bold=True, color="FFFFFF")
    VAL_FILL  = PatternFill("solid", fgColor="EBF3FB")
    thin      = Side(style="thin")
    BRD       = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr(ws, cols):
        for ci, (label, width) in enumerate(cols, 1):
            c = ws.cell(1, ci, label)
            c.fill = HDR_FILL; c.font = HDR_FONT
            c.alignment = Alignment(horizontal="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(ci)].width = width

    # ── Project sheet ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Project"
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 44
    for r, (lbl, val) in enumerate([
        ("Project Name",      os.path.splitext(os.path.basename(json_path))[0]),
        ("Address",           ""),
        ("Climate Zone",      climate_zone),
        ("Building Type",     "MultiFamily"),
        ("Front Orientation", front_orient),
        ("Standards Version", "2022"),
    ], start=2):
        ws.cell(r, 1, lbl).fill = HDR_FILL
        ws.cell(r, 1).font = HDR_FONT
        ws.cell(r, 2, val).fill = VAL_FILL

    # Check if multi-source (multiple drawings merged)
    multi = any('source' in z for z in zones)

    # ── Zones sheet ───────────────────────────────────────────────────────
    # Column order must match generate_gbxml.py expected input
    ws2 = wb.create_sheet("Zones")
    zone_cols = [
        ("Zone ID", 18), ("Zone Name", 26),
        ("Floor Area (sqft)", 18), ("Ceiling Height (ft)", 18),
        ("Condition Type", 26), ("Occupancy Type", 26),
    ]
    if multi:
        zone_cols.append(("Source Drawing", 30))
    hdr(ws2, zone_cols)
    for r, z in enumerate(zones, 2):
        vals = [
            z['id'], z.get('name', z['id']),
            round(z.get('area_sqft', 0), 1),
            z.get('ceiling_ht_ft', 9),
            z.get('condition', 'Conditioned'),
            z.get('occupancy', ''),
        ]
        if multi:
            vals.append(z.get('source', ''))
        for ci, val in enumerate(vals, 1):
            ws2.cell(r, ci, val)

    # ── Walls sheet ───────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Walls")
    wall_cols = [
        ("Wall ID", 20), ("Zone ID", 16), ("Wall Name", 22),
        ("Type", 28), ("Orientation", 16), ("Gross Area (sqft)", 18),
        ("Construction", 30), ("Adjacent Zone ID", 22),
    ]
    if multi:
        wall_cols.append(("Source Drawing", 30))
    hdr(ws3, wall_cols)
    for r, w in enumerate(walls, 2):
        vals = [
            w['wall_id'], w['zone_id'], w['name'],
            w['type'], w.get('orientation', ''),
            w['area_sqft'], '', w.get('adj_zone', ''),
        ]
        if multi:
            vals.append(w.get('source', ''))
        for ci, val in enumerate(vals, 1):
            ws3.cell(r, ci, val)

    # ── Openings sheet ────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Openings")
    open_cols = [
        ("Opening ID", 20), ("Wall ID", 20), ("Opening Name", 24),
        ("Type", 20), ("Area (sqft)", 14), ("U-Factor", 12), ("SHGC", 10),
    ]
    if multi:
        open_cols.append(("Source Drawing", 30))
    hdr(ws4, open_cols)
    for r, o in enumerate(openings, 2):
        vals = [
            o['id'], o['wall_id'], o['name'], o['type'],
            o['area_sqft'],
            o['ufactor'] if o['ufactor'] else '',
            o['shgc']    if o['shgc']    else '',
        ]
        if multi:
            vals.append(o.get('source', ''))
        for ci, val in enumerate(vals, 1):
            ws4.cell(r, ci, val)

    wb.save(out_path)
    return out_path


# ── Geometry JSON sidecar ─────────────────────────────────────────────────────

def write_geometry(json_path, zones, walls, openings):
    out_dir  = os.path.dirname(json_path)
    base     = os.path.splitext(os.path.basename(json_path))[0].replace('_t24', '')
    out_path = os.path.join(out_dir, base + '_t24_takeoff_geometry.json')

    geo = {'zones': {}, 'walls': {}, 'openings': {}}

    for z in zones:
        geo['zones'][z['id']] = {
            'vertices_ft':  pts_in_ft(z['vertices']),
            'centroid_ft':  [in2ft(z['centroid'][0]), in2ft(z['centroid'][1])],
            'floor':        z.get('floor', 1),
            'ceiling_ht_ft': z.get('ceiling_ht_ft', 9.0),
        }

    for w in walls:
        if w['p1'] is None:
            continue
        geo['walls'][w['wall_id']] = {
            'p1_ft': [in2ft(w['p1'][0]), in2ft(w['p1'][1])],
            'p2_ft': [in2ft(w['p2'][0]), in2ft(w['p2'][1])],
            'zone_id': w['zone_id'],
            'azimuth': round(w['azimuth'], 1),
        }

    for o in openings:
        pos = o.get('position', [0, 0])
        geo['openings'][o['id']] = {
            'position_ft': [in2ft(pos[0]), in2ft(pos[1])],
            'type':        o['type'],
            'wall_id':     o['wall_id'],
        }

    with open(out_path, 'w') as f:
        json.dump(geo, f, indent=2)
    return out_path


# ── Entry point ───────────────────────────────────────────────────────────────
def main():
    if len(sys.argv) < 2:
        print("Usage: python tz_to_excel.py <file.json> [file2.json ...]")
        print("       python tz_to_excel.py <directory>  (finds all *_t24.json)")
        sys.exit(1)

    # If a single directory is given, find all *_t24.json files in it
    if len(sys.argv) == 2 and os.path.isdir(sys.argv[1]):
        json_paths = sorted(glob.glob(os.path.join(sys.argv[1], "*_t24.json")))
        if not json_paths:
            print(f"No *_t24.json files found in: {sys.argv[1]}")
            sys.exit(1)
        print(f"Found {len(json_paths)} JSON files in {sys.argv[1]}")
    else:
        json_paths = sys.argv[1:]

    for p in json_paths:
        if not os.path.exists(p):
            print(f"File not found: {p}")
            sys.exit(1)

    # Process all files and merge
    all_zones = []
    all_walls = []
    all_openings = []
    climate_zone = ''
    front_orient = 'South'

    for jp in json_paths:
        print(f"Reading: {jp}")
        z, w, o, cz, fo = process(jp)
        dwg_name = os.path.splitext(os.path.basename(jp))[0].replace('_t24', '')

        # Tag each item with its source drawing
        for item in z:
            item['source'] = dwg_name
        for item in w:
            item['source'] = dwg_name
        for item in o:
            item['source'] = dwg_name

        all_zones.extend(z)
        all_walls.extend(w)
        all_openings.extend(o)
        # Use climate zone / orientation from first file
        if not climate_zone and cz:
            climate_zone = cz
            front_orient = fo

        print(f"  {len(z)} zones, {len(w)} wall segments, {len(o)} openings")

    print(f"\nTotal: {len(all_zones)} zones, {len(all_walls)} walls, {len(all_openings)} openings")

    # Use first file path for output naming
    primary_path = json_paths[0]
    if len(json_paths) > 1:
        # For multi-file, create a dummy path for naming only (file doesn't need to exist)
        out_dir = os.path.dirname(os.path.abspath(primary_path))
        primary_path = os.path.join(out_dir, "t24_combined_t24.json")

    xlsx = write_excel(primary_path, all_zones, all_walls, all_openings, climate_zone, front_orient)
    print(f"Excel:    {xlsx}")

    geo = write_geometry(primary_path, all_zones, all_walls, all_openings)
    print(f"Geometry: {geo}")
    print("Done. Review Excel, fill in Construction column, then run generate_gbxml.py.")

if __name__ == "__main__":
    main()
