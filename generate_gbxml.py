"""
generate_gbxml.py
Reads building data from an Excel template and generates a gbXML file
for import into EnergyPro 9.

Usage:
  python generate_gbxml.py                          <- creates blank template
  python generate_gbxml.py "My Project.xlsx"        <- generates gbXML from filled template
  python generate_gbxml.py "My Project.xlsx" out.xml
"""

import sys
import os
import math
import xml.etree.ElementTree as ET
from xml.dom import minidom
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# Orientation name -> azimuth degrees (clockwise from North)
ORIENTATION_MAP = {
    "north": 0,    "n": 0,
    "northeast": 45,  "ne": 45,
    "east": 90,    "e": 90,
    "southeast": 135, "se": 135,
    "south": 180,  "s": 180,
    "southwest": 225, "sw": 225,
    "west": 270,   "w": 270,
    "northwest": 315, "nw": 315,
    "front": 180,
    "back": 0,
    "left": 90,
    "right": 270,
}

SURFACE_TYPE_MAP = {
    "exterior wall": "ExteriorWall",
    "ext wall":      "ExteriorWall",
    "wall":          "ExteriorWall",
    "interior wall": "InteriorWall",
    "int wall":      "InteriorWall",
    "demising wall": "InteriorWall",
    "demising":      "InteriorWall",
    "roof":          "Roof",
    "ceiling":       "Ceiling",
    "interior ceiling": "InteriorCeiling",
    "exposed floor": "ExposedFloor",
    "slab":          "SlabOnGrade",
    "slab on grade": "SlabOnGrade",
    "raised floor":  "RaisedFloor",
    "underground wall": "UndergroundWall",
    "underground slab": "UndergroundSlab",
}

OPENING_TYPE_MAP = {
    "window":          "FixedWindow",
    "fixed window":    "FixedWindow",
    "operable window": "OperableWindow",
    "skylight":        "FixedSkylight",
    "fixed skylight":  "FixedSkylight",
    "operable skylight": "OperableSkylight",
    "door":            "NonSlidingDoor",
    "sliding door":    "SlidingDoor",
    "operable door":   "NonSlidingDoor",
}

CONDITION_TYPE_MAP = {
    "conditioned":       "HeatedAndCooled",
    "heated and cooled": "HeatedAndCooled",
    "heated only":       "HeatedOnly",
    "cooled only":       "CooledOnly",
    "unconditioned":     "Unconditioned",
    "semi-heated":       "SemiHeated",
    "semi heated":       "SemiHeated",
    "plenum":            "Plenum",
}

HORIZONTAL_SURFACES = {"Roof", "SlabOnGrade", "ExposedFloor", "RaisedFloor",
                       "UndergroundSlab", "Ceiling", "InteriorCeiling"}


# ---------------------------------------------------------------------------
# Excel template creator
# ---------------------------------------------------------------------------

def make_header(ws, cols, fill_color="1F4E79", font_color="FFFFFF"):
    hdr_fill = PatternFill("solid", fgColor=fill_color)
    hdr_font = Font(bold=True, color=font_color)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col_idx, (header, width) in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

def add_note_row(ws, text, col_count):
    r = ws.max_row + 1
    cell = ws.cell(row=r, column=1, value="# " + text)
    cell.font = Font(italic=True, color="595959")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=col_count)

def create_template(path: str):
    wb = Workbook()

    # ── Tab 1: Project ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Project"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 44

    info = [
        ("Project Name",      "My Building"),
        ("Address",           "123 Main St, Los Angeles, CA"),
        ("Climate Zone",      "CZ6  (Los Angeles)"),
        ("Building Type",     "MultiFamily"),
        ("Front Orientation", "South"),
        ("Standards Version", "2022"),
        ("Notes",             ""),
    ]
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    val_fill = PatternFill("solid", fgColor="EBF3FB")
    note_fill = PatternFill("solid", fgColor="FFF2CC")
    for r, (label, value) in enumerate(info, start=2):
        lc = ws.cell(row=r, column=1, value=label)
        lc.fill = hdr_fill
        lc.font = Font(bold=True, color="FFFFFF")
        vc = ws.cell(row=r, column=2, value=value)
        vc.fill = val_fill

    # Building type note
    note = ws.cell(row=10, column=1,
        value="# Building Type options: SingleFamily, MultiFamily, Office, RetailStore, "
              "Warehouse, Hotel, School, MedicalOffice, HighriseResidential, MidriseApartment")
    note.font = Font(italic=True, color="595959")
    ws.merge_cells("A10:B10")

    # ── Tab 2: Zones ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Zones")
    zone_cols = [
        ("Zone ID\n(unique, no spaces)",          20),
        ("Zone Name",                             28),
        ("Floor Area (sqft)",                     18),
        ("Ceiling Height (ft)",                   18),
        ("Condition Type\n(Conditioned / Unconditioned / Plenum)", 28),
        ("Occupancy Type\n(Residential / Office / Retail / etc.)", 28),
    ]
    make_header(ws2, zone_cols)
    sample_zones = [
        ("Z-101",   "Unit 101", 651,  8,  "Conditioned", "Residential"),
        ("Z-102",   "Unit 102", 900,  8,  "Conditioned", "Residential"),
        ("Z-LOBBY", "Lobby",    400,  10, "Conditioned", "Lobby"),
    ]
    for r, row in enumerate(sample_zones, start=2):
        for c, val in enumerate(row, start=1):
            ws2.cell(row=r, column=c, value=val)
    add_note_row(ws2,
        "Condition Type options: Conditioned, Heated Only, Cooled Only, Unconditioned, Semi-Heated, Plenum",
        6)
    add_note_row(ws2,
        "Zone ID must be unique and contain no spaces. Used to link walls and openings to zones.",
        6)

    # ── Tab 3: Walls ────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Walls")
    wall_cols = [
        ("Wall ID\n(unique, no spaces)",                           20),
        ("Zone ID",                                                16),
        ("Wall Name",                                              24),
        ("Type\n(Exterior Wall / Interior Wall / Roof / Slab on Grade)", 30),
        ("Orientation\n(N/S/E/W or degrees 0-360)",               22),
        ("Gross Area (sqft)\nincl. windows/doors",                 20),
        ("Construction",                                           30),
        ("Adjacent Zone ID\n(interior walls only)",                24),
    ]
    make_header(ws3, wall_cols)
    sample_walls = [
        ("W-101-N",  "Z-101", "North Wall",  "Exterior Wall", "North",  133, "R-21 Wood Framed Wall", ""),
        ("W-101-S",  "Z-101", "South Wall",  "Exterior Wall", "South",  133, "R-21 Wood Framed Wall", ""),
        ("W-101-E",  "Z-101", "East Wall",   "Exterior Wall", "East",   100, "R-21 Wood Framed Wall", ""),
        ("W-101-DM", "Z-101", "Demising",    "Interior Wall", "",       220, "R-0 Wall",              "Z-102"),
        ("W-101-RF", "Z-101", "Roof",        "Roof",          "",       651, "R-38 Roof",             ""),
        ("W-101-SL", "Z-101", "Slab",        "Slab on Grade", "",       651, "Slab-on-Grade",         ""),
    ]
    for r, row in enumerate(sample_walls, start=2):
        for c, val in enumerate(row, start=1):
            ws3.cell(row=r, column=c, value=val)
    add_note_row(ws3,
        "Gross Area = total surface area INCLUDING windows/doors/skylights. "
        "Orientation required for exterior walls (leave blank for roofs and slabs).",
        8)

    # ── Tab 4: Openings ─────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Openings")
    open_cols = [
        ("Opening ID\n(unique, no spaces)", 22),
        ("Wall ID\n(from Walls tab)",        18),
        ("Opening Name",                    24),
        ("Type\n(Window / Door / Skylight)", 22),
        ("Area (sqft)",                      14),
        ("U-Factor",                         12),
        ("SHGC",                             10),
    ]
    make_header(ws4, open_cols)
    sample_openings = [
        ("O-101-N-1",  "W-101-N",  "North Window A",  "Window",   27,   0.27, 0.18),
        ("O-101-N-2",  "W-101-N",  "North Window B",  "Window",   27,   0.27, 0.18),
        ("O-101-S-1",  "W-101-S",  "South Window A",  "Window",   48,   0.27, 0.18),
        ("O-101-E-1",  "W-101-E",  "East Door",       "Door",     24,   0.50, ""),
        ("O-101-SK-1", "W-101-RF", "Skylight A",      "Skylight", 12,   0.50, 0.25),
    ]
    for r, row in enumerate(sample_openings, start=2):
        for c, val in enumerate(row, start=1):
            ws4.cell(row=r, column=c, value=val)
    add_note_row(ws4,
        "Skylights reference a Roof Wall ID. U-Factor and SHGC required for windows and skylights; "
        "leave SHGC blank for doors.",
        7)

    wb.save(path)
    print(f"Template created: {path}")


# ---------------------------------------------------------------------------
# gbXML generator
# ---------------------------------------------------------------------------

def resolve_azimuth(orientation_str) -> float:
    if orientation_str in ("", None):
        return 0.0
    try:
        return float(orientation_str)
    except (ValueError, TypeError):
        return float(ORIENTATION_MAP.get(str(orientation_str).strip().lower(), 0))

def resolve_surface_type(type_str) -> str:
    return SURFACE_TYPE_MAP.get(str(type_str).strip().lower(), "ExteriorWall")

def resolve_opening_type(type_str) -> str:
    return OPENING_TYPE_MAP.get(str(type_str).strip().lower(), "FixedWindow")

def resolve_condition_type(type_str) -> str:
    return CONDITION_TYPE_MAP.get(str(type_str).strip().lower(), "HeatedAndCooled")

def tilt_for_surface(surface_type: str) -> float:
    return 0.0 if surface_type in HORIZONTAL_SURFACES else 90.0

def add_rect_geometry(parent, azimuth: float, tilt: float, width: float, height: float):
    """Add RectangularGeometry child — this is what EnergyPro reads for area and orientation."""
    rg = ET.SubElement(parent, "RectangularGeometry")
    ET.SubElement(rg, "Azimuth").text  = f"{azimuth:.4f}"
    cp = ET.SubElement(rg, "CartesianPoint")
    for coord in ("0", "0", "0"):
        ET.SubElement(cp, "Coordinate").text = coord
    ET.SubElement(rg, "Width").text  = f"{width:.4f}"
    ET.SubElement(rg, "Height").text = f"{height:.4f}"
    ET.SubElement(rg, "Tilt").text   = f"{tilt:.4f}"


def generate_gbxml(xlsx_path: str, out_path: str):
    wb = openpyxl.load_workbook(xlsx_path)

    # -- Project info --
    ws_proj = wb["Project"]
    def proj(r): return str(ws_proj.cell(row=r, column=2).value or "")

    project_name  = proj(2)
    address       = proj(3)
    building_type = proj(5) or "MultiFamily"

    # -- Zones --
    ws_zones = wb["Zones"]
    zones = []
    zone_height = {}  # zone_id -> ceiling height (for wall geometry calc)
    for row in ws_zones.iter_rows(min_row=2, values_only=True):
        if not row[0] or str(row[0]).startswith("#"): continue
        zid, name, area, height, ctype, occ = (list(row) + [None]*6)[:6]
        if not zid: continue
        zid = str(zid).strip().replace(" ", "_")
        h = float(height or 9)
        zone_height[zid] = h
        zones.append({
            "id":        zid,
            "name":      str(name or zid),
            "area":      float(area or 0),
            "height":    h,
            "cond_type": resolve_condition_type(ctype or "Conditioned"),
            "occ_type":  str(occ or "").strip(),
        })

    # -- Walls --
    ws_walls = wb["Walls"]
    walls = []
    for row in ws_walls.iter_rows(min_row=2, values_only=True):
        if not row[0] or str(row[0]).startswith("#"): continue
        wid, zid, name, wtype, orient, area, construction, adj_zone = (list(row) + [None]*8)[:8]
        if not wid: continue
        wid  = str(wid).strip().replace(" ", "_")
        zid  = str(zid or "").strip().replace(" ", "_")
        stype = resolve_surface_type(wtype or "Exterior Wall")
        gross_area = float(area or 0)
        h = zone_height.get(zid, 9.0)

        # Rectangular geometry dimensions
        if stype in HORIZONTAL_SURFACES:
            # Flat surface — use square approximation
            side = math.sqrt(gross_area) if gross_area > 0 else 1.0
            rg_width, rg_height = side, side
        else:
            # Vertical wall — height = zone ceiling height, width = area / height
            rg_height = h
            rg_width  = (gross_area / h) if h > 0 else gross_area

        walls.append({
            "id":           wid,
            "zone_id":      zid,
            "name":         str(name or wid),
            "surface_type": stype,
            "azimuth":      resolve_azimuth(orient),
            "tilt":         tilt_for_surface(stype),
            "area":         gross_area,
            "rg_width":     rg_width,
            "rg_height":    rg_height,
            "construction": str(construction or ""),
            "adj_zone":     str(adj_zone or "").strip().replace(" ", "_"),
        })

    # -- Openings --
    ws_open = wb["Openings"]
    openings = []
    for row in ws_open.iter_rows(min_row=2, values_only=True):
        if not row[0] or str(row[0]).startswith("#"): continue
        oid, wall_id, name, otype, area, ufactor, shgc = (list(row) + [None]*7)[:7]
        if not oid: continue
        o_area = float(area or 0)
        side   = math.sqrt(o_area) if o_area > 0 else 1.0
        openings.append({
            "id":       str(oid).strip().replace(" ", "_"),
            "wall_id":  str(wall_id or "").strip().replace(" ", "_"),
            "name":     str(name or oid),
            "type":     resolve_opening_type(otype or "Window"),
            "area":     o_area,
            "rg_width": side,
            "rg_height":side,
            "ufactor":  float(ufactor) if ufactor not in ("", None) else None,
            "shgc":     float(shgc)    if shgc    not in ("", None) else None,
        })

    openings_by_wall = {}
    for o in openings:
        openings_by_wall.setdefault(o["wall_id"], []).append(o)

    # -- Build gbXML tree --
    root = ET.Element("gbXML", {
        "xmlns":           "http://www.gbxml.org/schema",
        "temperatureUnit": "F",
        "lengthUnit":      "Feet",
        "areaUnit":        "SquareFeet",
        "volumeUnit":      "CubicFeet",
        "version":         "6.01",
    })

    campus = ET.SubElement(root, "Campus", {"id": "campus-1"})
    ET.SubElement(campus, "Name").text     = project_name
    ET.SubElement(campus, "Location").text = address

    total_area = sum(z["area"] for z in zones)
    building = ET.SubElement(campus, "Building", {
        "id":           "building-1",
        "buildingType": building_type,
    })
    ET.SubElement(building, "Name").text = project_name
    ET.SubElement(building, "Area").text = str(total_area)

    # Spaces
    for z in zones:
        attrs = {
            "id":           z["id"],
            "zoneIdRef":    z["id"],
            "conditionType": z["cond_type"],
        }
        if z["occ_type"]:
            attrs["occupancyType"] = z["occ_type"]
        space = ET.SubElement(building, "Space", attrs)
        ET.SubElement(space, "Name").text         = z["name"]
        ET.SubElement(space, "Area").text         = str(z["area"])
        ET.SubElement(space, "Volume").text       = str(round(z["area"] * z["height"], 2))
        ET.SubElement(space, "CeilingHeight").text = str(z["height"])

    # Surfaces (at Campus level)
    for w in walls:
        surf = ET.SubElement(campus, "Surface", {
            "id":          w["id"],
            "surfaceType": w["surface_type"],
        })
        ET.SubElement(surf, "Name").text = w["name"]

        if w["construction"]:
            ET.SubElement(surf, "CADObjectId").text = w["construction"]

        if w["zone_id"]:
            ET.SubElement(surf, "AdjacentSpaceId", {"spaceIdRef": w["zone_id"]})
        if w["adj_zone"]:
            ET.SubElement(surf, "AdjacentSpaceId", {"spaceIdRef": w["adj_zone"]})

        # RectangularGeometry — EnergyPro reads area and orientation from here
        add_rect_geometry(surf, w["azimuth"], w["tilt"], w["rg_width"], w["rg_height"])

        # Openings
        for o in openings_by_wall.get(w["id"], []):
            opening = ET.SubElement(surf, "Opening", {
                "id":          o["id"],
                "openingType": o["type"],
            })
            ET.SubElement(opening, "Name").text = o["name"]
            if o["ufactor"] is not None:
                ET.SubElement(opening, "U-value").text = str(o["ufactor"])
            if o["shgc"] is not None:
                ET.SubElement(opening, "SHGC").text = str(o["shgc"])
            add_rect_geometry(opening, w["azimuth"], w["tilt"], o["rg_width"], o["rg_height"])

    # -- Write pretty XML --
    rough  = ET.tostring(root, encoding="unicode")
    rough  = rough.replace(' xmlns=""', '').replace(" xmlns=''", '')
    pretty = minidom.parseString(rough).toprettyxml(indent="  ")
    lines  = pretty.split("\n")
    if lines[0].startswith("<?xml"):
        lines[0] = '<?xml version="1.0" encoding="utf-8"?>'
    output = "\n".join(lines)
    output = output.replace(' xmlns=""', '').replace(" xmlns=''", '')

    with open(out_path, "w", encoding="utf-8") as f:
        f.write(output)

    print(f"gbXML generated: {out_path}")
    print(f"  {len(zones)} zone(s), {len(walls)} surface(s), {len(openings)} opening(s)")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main():
    args = sys.argv[1:]
    if not args:
        template_path = os.path.join(SCRIPT_DIR, "T24 Input Template.xlsx")
        create_template(template_path)
        print("\nFill in the template, then run:")
        print('  python generate_gbxml.py "T24 Input Template.xlsx"')
        return

    xlsx_path = args[0]
    if not os.path.isabs(xlsx_path):
        xlsx_path = os.path.join(SCRIPT_DIR, xlsx_path)
    if not os.path.exists(xlsx_path):
        print(f"File not found: {xlsx_path}")
        sys.exit(1)

    out_path = args[1] if len(args) >= 2 else os.path.splitext(xlsx_path)[0] + ".xml"
    generate_gbxml(xlsx_path, out_path)


if __name__ == "__main__":
    main()
